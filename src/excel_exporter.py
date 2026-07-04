"""
excel_exporter.py
Export SalesOrder to Excel.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from app_config import OUTPUT_DIR
from logger import logger
from models import SalesOrder


class ExcelExporter:
    """
    Export SalesOrder objects to Excel.
    """

    HEADERS = [
        "Customer",
        "PO Number",
        "Material",
        "Customer Material",
        "Description",
        "Quantity",
        "Unit",
        "Unit Price",
        "Amount",
        "Delivery Date",
    ]

    def export(
        self,
        order: SalesOrder,
        output_file: Path | None = None,
    ) -> Path:

        if output_file is None:
            filename = (
                order.po_no
                if order.po_no
                else "SalesOrder"
            )
            output_file = OUTPUT_DIR / f"{filename}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Order"

        # Header
        for col, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        # Data
        row = 2

        for item in order.items:

            ws.cell(row=row, column=1).value = order.customer
            ws.cell(row=row, column=2).value = order.po_no
            ws.cell(row=row, column=3).value = item.material_no
            ws.cell(row=row, column=4).value = item.customer_material
            ws.cell(row=row, column=5).value = item.description
            ws.cell(row=row, column=6).value = item.quantity
            ws.cell(row=row, column=7).value = item.unit
            ws.cell(row=row, column=8).value = item.unit_price
            ws.cell(row=row, column=9).value = item.amount
            ws.cell(row=row, column=10).value = item.delivery_date

            row += 1

        wb.save(output_file)

        logger.info(
            "Excel exported: %s",
            output_file.name,
        )

        return output_file


if __name__ == "__main__":

    exporter = ExcelExporter()

    order = SalesOrder(
        customer="ABB",
        po_no="4500012345",
    )

    exporter.export(order)