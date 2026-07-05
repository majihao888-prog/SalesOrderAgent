"""
Summary Excel Service
"""

from __future__ import annotations

from pathlib import Path
from datetime import datetime

from openpyxl import Workbook, load_workbook

from app_config import OUTPUT_DIR
from models import SalesOrder


class SummaryService:

    HEADERS = [

        "Customer",
        "PO",

        "Order Date",
        "Currency",

        "Material",
        "Description",

        "Qty",
        "Unit",

        "Unit Price",
        "Amount",

        "Delivery Date",

        "Status",

        "PDF",

        "Process Time",
    ]

    def append(
        self,
        order: SalesOrder,
        pdf: Path,
    ) -> None:

        year = "UNKNOWN"

        if order.order_date:

            try:

                year = order.order_date.split(".")[-1]

            except Exception:

                pass

        summary_dir = OUTPUT_DIR / "summary"

        summary_dir.mkdir(
            parents=True,
            exist_ok=True,
        )

        excel = summary_dir / f"{year}.xlsx"

        # ----------------------------
        # Create workbook
        # ----------------------------

        if excel.exists():

            wb = load_workbook(excel)

            ws = wb.active

        else:

            wb = Workbook()

            ws = wb.active

            ws.title = year

            ws.append(self.HEADERS)

        # ----------------------------
        # Write Items
        # ----------------------------

        if not order.items:

            ws.append([
                order.customer,
                order.po_no,
                order.order_date,
                order.currency,
                "",
                "",
                "",
                "",
                "",
                order.total_amount,
                "",
                "SUCCESS",
                pdf.name,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ])

        else:

            for item in order.items:

                ws.append([

                    order.customer,

                    order.po_no,

                    order.order_date,

                    order.currency,

                    item.material_no,

                    item.description,

                    item.quantity,

                    item.unit,

                    item.unit_price,

                    item.amount,

                    item.delivery_date,

                    "SUCCESS",

                    pdf.name,

                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),

                ])

        wb.save(excel)